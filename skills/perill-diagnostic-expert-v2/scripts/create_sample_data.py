#!/usr/bin/env python3
"""Create sample PERILL test data Excel file"""

import openpyxl
from openpyxl import Workbook
import random

# PERILL dimensions and questions
DIMENSIONS = {
    'P': {'name': 'Purpose and motivation\n目的与动机', 'questions': [
        'Team members clearly understand the team mission and goals\n团队成员清楚理解团队使命和目标',
        'Team goals are aligned with organizational strategy\n团队目标与组织战略一致',
        'Team members share common values and beliefs\n团队成员拥有共同的价值观',
        'Each member understands their role in achieving team goals\n每个成员都理解自己在实现团队目标中的角色',
        'The team has a clear and inspiring vision\n团队有清晰且具激励性的愿景',
        'Team purpose motivates members daily\n团队使命每天激励着成员'
    ]},
    'E': {'name': 'External system and process\n外部系统与流程', 'questions': [
        'The team effectively communicates with external stakeholders\n团队与外部利益相关方有效沟通',
        'The team responds timely to market changes\n团队及时响应市场变化',
        'The team has good relationships with partner organizations\n团队与合作伙伴关系良好',
        'The team understands customer needs and expectations\n团队了解客户需求和期望',
        'The team accesses necessary resources from external environment\n团队从外部环境获取所需资源',
        'The team maintains good reputation in the industry\n团队在行业中保持良好声誉'
    ]},
    'R': {'name': 'Relationship\n关系', 'questions': [
        'Team members trust each other\n团队成员相互信任',
        'Team members can express opinions freely without fear\n团队成员可以自由表达意见而不担心',
        'Conflicts are handled constructively within the team\n团队内冲突被建设性处理',
        'Team members support each other in difficult times\n团队成员在困难时相互支持',
        'There is strong cohesion among team members\n团队成员间有很强的凝聚力',
        'Team members appreciate each other\'s contributions\n团队成员相互欣赏彼此的贡献'
    ]},
    'I': {'name': 'Internal system and process\n内部系统与流程', 'questions': [
        'Team workflows and processes are clearly defined\n团队工作流程清晰定义',
        'Decision-making is efficient and effective\n决策高效且有效',
        'Roles and responsibilities are clearly divided\n角色和职责分工明确',
        'Information flows smoothly within the team\n团队内信息流通顺畅',
        'The team has effective meeting management\n团队有有效的会议管理',
        'The team continuously improves its processes\n团队持续改进流程'
    ]},
    'L': {'name': 'Learning\n学习', 'questions': [
        'The team regularly reflects on its performance\n团队定期反思绩效',
        'Team members have personal development plans\n团队成员有个人发展计划',
        'The team learns from both successes and failures\n团队从成功和失败中学习',
        'Knowledge sharing is encouraged and practiced\n知识共享被鼓励和实践',
        'The team adopts new methods and tools\n团队采用新方法和工具',
        'Lessons learned are documented and applied\n经验教训被记录和应用'
    ]},
    'Ld': {'name': 'Leadership\n领导力', 'questions': [
        'Leadership is distributed across the team\n领导力在团队中分布',
        'The leader empowers team members effectively\n领导者有效授权团队成员',
        'The leader provides clear direction and guidance\n领导者提供清晰的方向和指导',
        'The leader fosters a positive team culture\n领导者培育积极的团队文化',
        'The leader handles conflicts and challenges well\n领导者处理冲突和挑战得当',
        'The leader develops future leaders within the team\n领导者在团队中培养未来领导者'
    ]}
}

DIMENSION_ORDER = ['P', 'E', 'R', 'I', 'L', 'Ld']


def create_sample_excel(output_path: str):
    """Create a sample PERILL Excel file"""

    wb = Workbook()
    ws = wb.active
    ws.title = "PERILL Data"

    # Row 1: Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'PERILL Team Diagnostic Questionnaire / PERILL团队诊断问卷'
    ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)

    # Row 2: Headers
    ws['A2'] = 'Dimension / 维度'
    ws['B2'] = 'Question / 问题'
    ws['C2'] = 'Type / 类型'
    ws['D2'] = 'Recorder / 记录人'
    ws['E2'] = 'Team Member 1\n团队成员1'
    ws['F2'] = 'Team Member 2\n团队成员2'
    ws['G2'] = 'Team Member 3\n团队成员3'
    ws['H2'] = 'Team Leader\n团队领导'
    ws['I2'] = 'Stakeholder 1\n相关方1'
    ws['J2'] = 'Stakeholder 2\n相关方2'

    # Row 3: Roles
    ws['A3'] = ''
    ws['B3'] = ''
    ws['C3'] = ''
    ws['D3'] = ''
    ws['E3'] = 'Team Member'
    ws['F3'] = 'Team Member'
    ws['G3'] = 'Team Member'
    ws['H3'] = 'Team Leader'
    ws['I3'] = 'Stakeholder'
    ws['J3'] = 'Stakeholder'

    # Make header row bold
    for cell in ws[2]:
        if cell.value:
            cell.font = openpyxl.styles.Font(bold=True)

    # Data rows starting from row 4
    row = 4
    q_num = 1

    # Set random seed for reproducibility
    random.seed(42)

    for dim in DIMENSION_ORDER:
        dim_info = DIMENSIONS[dim]
        questions = dim_info['questions']

        for i, q_text in enumerate(questions):
            # Current score row
            ws.cell(row=row, column=1, value=f"{dim} - {dim_info['name']}")
            ws.cell(row=row, column=2, value=q_text)
            ws.cell(row=row, column=3, value='Current / 现状')
            ws.cell(row=row, column=4, value=f'Recorder {q_num}')

            # Generate realistic scores with some variation
            # Team members tend to score higher than reality
            base_team = random.uniform(2.5, 4.0)
            base_stakeholder = random.uniform(2.0, 3.5)

            for col in range(5, 11):  # Columns E-J
                if col <= 8:  # Team members/leader
                    score = round(random.uniform(max(1, base_team - 0.5), min(5, base_team + 0.5)), 1)
                else:  # Stakeholders
                    score = round(random.uniform(max(1, base_stakeholder - 0.3), min(5, base_stakeholder + 0.3)), 1)
                ws.cell(row=row, column=col, value=score)

            row += 1

            # Expected score row
            ws.cell(row=row, column=1, value=f"{dim} - {dim_info['name']}")
            ws.cell(row=row, column=2, value=q_text)
            ws.cell(row=row, column=3, value='Expected / 期望')
            ws.cell(row=row, column=4, value=f'Recorder {q_num}')

            # Expected scores are typically higher
            for col in range(5, 11):
                if col <= 8:
                    score = round(min(5, random.uniform(base_team + 0.8, base_team + 1.5)), 1)
                else:
                    score = round(min(5, random.uniform(base_stakeholder + 0.5, base_stakeholder + 1.2)), 1)
                ws.cell(row=row, column=col, value=score)

            row += 1
            q_num += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    for col in 'EFGHIJ':
        ws.column_dimensions[col].width = 15

    wb.save(output_path)
    print(f"Sample PERILL data created: {output_path}")


if __name__ == '__main__':
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else 'sample_perill_data.xlsx'
    create_sample_excel(output)